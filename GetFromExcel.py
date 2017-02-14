# Import modules
import openpyxl
import os
filename = ""
filename_dir = ""
current_dir = ""

# Function to gather information about the current working directory and file directory


def get_current_dir():
    global filename
    filename = "twInventory7.xlsx"
    global filename_dir
    filename_dir = r"C:\CSI"
    global current_dir
    current_dir = os.getcwd()
    print("The current working directory is %s " % current_dir)

# Function to set the current working directory to the file directory if they are not the same


def get_set_working_dir():
    if current_dir != filename_dir:
        os.chdir(filename_dir)
        working_dir = os.getcwd()
        print("Changed the current directory to match the directory where the file is located: %s " % working_dir)
    else:
        print("Current directory matches the directory where the file is located: %s " % current_dir)

# Function to read the information from the top columns and the rest of the excel and put them in a dictionary


def read_excel():
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name("windowsClusterInfo")
    excel_obj = {}
    col_arr = []
    top_col = sheet.max_column
    top_row = sheet.max_row

    for col in range(1, top_col + 1):
        col_arr.append(sheet.cell(row=1, column=col).value)

    for row_val in range(2, top_row + 1):
        for col_val in range(1, top_col + 1):

            excel_obj[col_arr[col_val - 1] + str(row_val - 1)] = sheet.cell(row=row_val, column=col_val).value
    print(excel_obj)


# Main function

           


def main():
    get_current_dir()
    get_set_working_dir()
    read_excel()

# Call main function
main()
